from django.shortcuts import render
from django.http import HttpResponse
from django.db.models import Sum, Count
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, Alignment
import io

from sales.models import Invoice, InvoiceItem
from inventory.models import StockMovement


def reports_dashboard(request):
    """Reports dashboard"""
    # Calculate current month stats
    now = timezone.now()
    start_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    # Monthly earnings (total sales)
    monthly_earnings = Invoice.objects.filter(
        date__gte=start_of_month.date()
    ).aggregate(total=Sum('total_incl'))['total'] or Decimal('0')
    
    # Monthly spending (stock purchases)
    monthly_spending = StockMovement.objects.filter(
        reason='purchase',
        timestamp__gte=start_of_month
    ).aggregate(
        total=Sum('qty_change') * Sum('unit_cost')
    )
    
    # Calculate spending manually due to aggregation complexity
    purchases = StockMovement.objects.filter(
        reason='purchase',
        timestamp__gte=start_of_month,
        unit_cost__isnull=False
    )
    total_spending = sum(
        (movement.qty_change * movement.unit_cost) 
        for movement in purchases
    ) or Decimal('0')
    
    # Sales count
    monthly_sales_count = Invoice.objects.filter(
        date__gte=start_of_month.date()
    ).count()
    
    # Recent invoices
    recent_invoices = Invoice.objects.order_by('-created_at')[:10]
    
    monthly_profit = monthly_earnings - total_spending
    
    context = {
        'monthly_earnings': monthly_earnings,
        'monthly_spending': total_spending,
        'monthly_profit': monthly_profit,
        'monthly_sales_count': monthly_sales_count,
        'recent_invoices': recent_invoices,
        'current_month': now.strftime('%B %Y'),
        'page_title': 'Reports Dashboard'
    }
    return render(request, 'reports/dashboard.html', context)


def monthly_summary(request):
    """Monthly summary report"""
    year = int(request.GET.get('year', timezone.now().year))
    month = int(request.GET.get('month', timezone.now().month))
    
    # Create date range for the month
    start_date = datetime(year, month, 1)
    if month == 12:
        end_date = datetime(year + 1, 1, 1)
    else:
        end_date = datetime(year, month + 1, 1)
    
    # Get monthly data
    invoices = Invoice.objects.filter(
        date__gte=start_date.date(),
        date__lt=end_date.date()
    )
    
    purchases = StockMovement.objects.filter(
        reason='purchase',
        timestamp__gte=start_date,
        timestamp__lt=end_date,
        unit_cost__isnull=False
    )
    
    total_earnings = invoices.aggregate(total=Sum('total_incl'))['total'] or Decimal('0')
    total_spending = sum(
        (movement.qty_change * movement.unit_cost) 
        for movement in purchases
    ) or Decimal('0')
    
    # Prepare month names for template
    month_names = [
        (1, 'January'), (2, 'February'), (3, 'March'), (4, 'April'),
        (5, 'May'), (6, 'June'), (7, 'July'), (8, 'August'),
        (9, 'September'), (10, 'October'), (11, 'November'), (12, 'December')
    ]
    
    context = {
        'year': year,
        'month': month,
        'month_name': start_date.strftime('%B'),
        'month_names': month_names,
        'total_earnings': total_earnings,
        'total_spending': total_spending,
        'profit': total_earnings - total_spending,
        'invoice_count': invoices.count(),
        'invoices': invoices.order_by('-date'),
        'page_title': f'Monthly Summary - {start_date.strftime("%B %Y")}'
    }
    return render(request, 'reports/monthly_summary.html', context)


def sales_report(request):
    """Sales report"""
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    invoices = Invoice.objects.all()
    
    if start_date:
        invoices = invoices.filter(date__gte=start_date)
    if end_date:
        invoices = invoices.filter(date__lte=end_date)
    
    invoices = invoices.order_by('-date')
    
    total_sales = invoices.aggregate(total=Sum('total_incl'))['total'] or Decimal('0')
    total_base = invoices.aggregate(total=Sum('total_base'))['total'] or Decimal('0')
    total_gst = invoices.aggregate(total=Sum('total_tax'))['total'] or Decimal('0')
    average_sale = total_sales / invoices.count() if invoices.count() > 0 else Decimal('0')
    
    context = {
        'invoices': invoices,
        'total_sales': total_sales,
        'total_base': total_base,
        'total_gst': total_gst,
        'average_sale': average_sale,
        'start_date': start_date,
        'end_date': end_date,
        'page_title': 'Sales Report'
    }
    return render(request, 'reports/sales_report.html', context)


def stock_report(request):
    """Stock movement report"""
    movements = StockMovement.objects.select_related('product').order_by('-timestamp')[:100]
    
    context = {
        'movements': movements,
        'page_title': 'Stock Report'
    }
    return render(request, 'reports/stock_report.html', context)


def export_sales(request, format):
    """Export sales data"""
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    invoices = Invoice.objects.all()
    
    if start_date:
        invoices = invoices.filter(date__gte=start_date)
    if end_date:
        invoices = invoices.filter(date__lte=end_date)
    
    if format == 'xlsx':
        return export_sales_xlsx(invoices)
    else:
        return HttpResponse('Invalid format', status=400)


def export_monthly(request, format):
    """Export monthly summary"""
    year = int(request.GET.get('year', timezone.now().year))
    month = int(request.GET.get('month', timezone.now().month))
    
    start_date = datetime(year, month, 1)
    if month == 12:
        end_date = datetime(year + 1, 1, 1)
    else:
        end_date = datetime(year, month + 1, 1)
    
    invoices = Invoice.objects.filter(
        date__gte=start_date.date(),
        date__lt=end_date.date()
    )
    
    if format == 'xlsx':
        return export_monthly_xlsx(invoices, start_date)
    else:
        return HttpResponse('Invalid format', status=400)


def export_sales_xlsx(invoices):
    """Export sales to Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Report"
    
    # Headers
    headers = ['Invoice Number', 'Date', 'Customer', 'Total (Incl Tax)', 'Total (Base)', 'Tax Amount', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for row, invoice in enumerate(invoices, 2):
        ws.cell(row=row, column=1, value=invoice.number)
        ws.cell(row=row, column=2, value=invoice.date)
        ws.cell(row=row, column=3, value=invoice.get_customer_display_name())
        ws.cell(row=row, column=4, value=float(invoice.total_incl))
        ws.cell(row=row, column=5, value=float(invoice.total_base))
        ws.cell(row=row, column=6, value=float(invoice.total_tax))
        ws.cell(row=row, column=7, value=invoice.payment_status)
    
    # Save to memory
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="sales_report.xlsx"'
    return response


def export_monthly_xlsx(invoices, start_date):
    """Export monthly summary to Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Monthly Summary - {start_date.strftime('%B %Y')}"
    
    # Title
    ws.cell(row=1, column=1, value=f"Monthly Summary - {start_date.strftime('%B %Y')}")
    ws.cell(row=1, column=1).font = Font(size=16, bold=True)
    
    # Headers
    headers = ['Invoice Number', 'Date', 'Customer', 'Total Amount', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
    
    # Data
    for row, invoice in enumerate(invoices, 4):
        ws.cell(row=row, column=1, value=invoice.number)
        ws.cell(row=row, column=2, value=invoice.date)
        ws.cell(row=row, column=3, value=invoice.get_customer_display_name())
        ws.cell(row=row, column=4, value=float(invoice.total_incl))
        ws.cell(row=row, column=5, value=invoice.payment_status)
    
    # Summary
    total = sum(float(inv.total_incl) for inv in invoices)
    summary_row = len(invoices) + 5
    ws.cell(row=summary_row, column=3, value="Total:")
    ws.cell(row=summary_row, column=3).font = Font(bold=True)
    ws.cell(row=summary_row, column=4, value=total)
    ws.cell(row=summary_row, column=4).font = Font(bold=True)
    
    # Save to memory
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="monthly_summary_{start_date.strftime("%Y_%m")}.xlsx"'
    return response